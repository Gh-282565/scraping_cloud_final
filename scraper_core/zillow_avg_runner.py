# -*- coding: utf-8 -*-
import sys, os, json, re, traceback, glob
from urllib.parse import quote
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill

LOG_FILE = "runner_debug.log"
OUT_BASE = "risultati_zillow_media.xlsx"
SQFT_PER_ACRE = 43560.0

def log(*a):
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(" ".join(str(x) for x in a) + "\n")
    except Exception:
        pass
    print(*a, flush=True)

# ------------------------------
# Import scraper
# ------------------------------
try:
    from . import zillow_test_scrape as zts
    log("[INIT] zillow_test_scrape import OK")
except Exception as e:
    log("[ERR] Impossibile importare zillow_test_scrape:", e)
    log(traceback.format_exc())
    sys.exit(3)

US_STATE_NAMES = {
    "AL":"Alabama","AK":"Alaska","AZ":"Arizona","AR":"Arkansas","CA":"California","CO":"Colorado","CT":"Connecticut",
    "DE":"Delaware","FL":"Florida","GA":"Georgia","HI":"Hawaii","ID":"Idaho","IL":"Illinois","IN":"Indiana","IA":"Iowa",
    "KS":"Kansas","KY":"Kentucky","LA":"Louisiana","ME":"Maine","MD":"Maryland","MA":"Massachusetts","MI":"Michigan",
    "MN":"Minnesota","MS":"Mississippi","MO":"Missouri","MT":"Montana","NE":"Nebraska","NV":"Nevada","NH":"New Hampshire",
    "NJ":"New Jersey","NM":"New Mexico","NY":"New York","NC":"North Carolina","ND":"North Dakota","OH":"Ohio","OK":"Oklahoma",
    "OR":"Oregon","PA":"Pennsylvania","RI":"Rhode Island","SC":"South Carolina","SD":"South Dakota","TN":"Tennessee",
    "TX":"Texas","UT":"Utah","VT":"Vermont","VA":"Virginia","WA":"Washington","WV":"West Virginia","WI":"Wisconsin","WY":"Wyoming",
    "DC":"District of Columbia","PR":"Puerto Rico"
}

def state_full_name(s: str) -> str:
    if not s:
        return ""
    code = str(s).strip().upper()
    return US_STATE_NAMES.get(code, s)

def _lot_bounds(min_lot, max_lot):
    def to_int(x):
        if x is None: return None
        s = str(x).strip()
        if s.lower().startswith("no "): return None
        try:
            return int(float(s.replace(",", ".")))
        except Exception:
            return None
    return to_int(min_lot), to_int(max_lot)

def build_url(contea, stato, region_id, north, south, east, west, period, min_lot, max_lot, tipo_vendita="land"):
    min_lot_i, max_lot_i = _lot_bounds(min_lot, max_lot)
    if min_lot_i is not None: min_lot_i = int(min_lot_i * SQFT_PER_ACRE)
    if max_lot_i is not None: max_lot_i = int(max_lot_i * SQFT_PER_ACRE)

    filter_state = {
        "sort": {"value": "globalrelevanceex"},
        "sf":   {"value": False}, "tow":  {"value": False}, "mf":   {"value": False}, "con":  {"value": False},
        "apa":  {"value": False}, "manu": {"value": False}, "apco": {"value": False},
        "lot":  {"min": min_lot_i, "max": max_lot_i, "units": None},
        "doz":  {"value": f"{period}m"} if str(period).isdigit() else {"value": str(period)},
        "land": {"value": True}
    }
    if str(tipo_vendita).lower() == "sold":
        filter_state.update({
            "rs":   {"value": True},
            "fsba": {"value": False},
            "fsbo": {"value": False},
            "nc":   {"value": False},
            "cmsn": {"value": False},
            "auc":  {"value": False},
            "fore": {"value": False},
        })

    query = {
        "pagination": {}, "isMapVisible": True,
        "mapBounds": {"west": float(west), "east": float(east), "south": float(south), "north": float(north)} if all([west,east,south,north]) else None,
        "regionSelection": [{"regionId": int(region_id), "regionType": 4}] if str(region_id).strip().isdigit() else [],
        "filterState": filter_state, "isListVisible": True, "usersSearchTerm": f"{contea} County {stato}"
    }
    query = {k:v for k,v in query.items() if v is not None}
    qs = quote(json.dumps(query, separators=(",",":")))
    tipo_segment = "sold" if str(tipo_vendita).lower() == "sold" else "land"
    return f"https://www.zillow.com/{str(contea).lower().replace(' ','-')}-county-{str(stato).lower()}/{tipo_segment}/?searchQueryState={qs}"

def _to_float(x):
    try:
        return float(str(x).replace(",",""))
    except Exception:
        return None

_acres_pat = re.compile(r"([\d.,]+)\s*(?:acres?|ac|Acre|Acres)\b", re.IGNORECASE)
_sqft_pat  = re.compile(r"([\d.,]+)\s*(?:sq\s*ft|sqft|square\s*feet)\b", re.IGNORECASE)

def _parse_acres_fallback(r):
    for field in ["acres","lot_size","meta","details","subtitle","description","info"]:
        val = getattr(r, field, None)
        if not val: continue
        s = str(val)
        m = _acres_pat.search(s)
        if m:
            v = _to_float(m.group(1))
            if v is not None: return v
        m = _sqft_pat.search(s)
        if m:
            v = _to_float(m.group(1))
            if v is not None: return v / SQFT_PER_ACRE
    return None

def df_from_rows(rows):
    data = []
    for r in rows:
        price_txt = getattr(r,"price",None)
        price_num = None
        if price_txt:
            s = re.sub(r"[^0-9.\-]", "", str(price_txt))
            price_num = float(s) if s else None
        acres = getattr(r,"acres",None) or _parse_acres_fallback(r)
        acres_num = None
        if acres not in ("",None):
            try: acres_num = float(str(acres).replace(",",""))
            except: acres_num = None
        ppa = price_num / acres_num if (price_num and acres_num and acres_num>0) else None
        data.append({
            "Price":price_txt,
            "Price_num":price_num,
            "Acres":acres,
            "Acres_num":acres_num,
            "Price_per_Acre":ppa,
            "Location":getattr(r,"location",None),
            "Link":getattr(r,"link",None)
        })
    return pd.DataFrame(data, columns=["Price","Price_num","Acres","Acres_num","Price_per_Acre","Location","Link"])

def _auto_fit(ws):
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, ws.max_column+1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
            for cell in row:
                val = cell.value
                l = len(str(val)) if val is not None else 0
                if l > max_len: max_len = l
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len+2, 80)

def append_sheet_with_avg(book_path, sheet_name, df, stato, contea):
    from openpyxl.styles import numbers, Border, Side
    # stats
    avg_price = float(df["Price_num"].dropna().mean()) if "Price_num" in df.columns and not df.empty else None
    avg_ppa   = float(df["Price_per_Acre"].dropna().mean()) if "Price_per_Acre" in df.columns and not df.empty else None
    med_price = float(df["Price_num"].dropna().median()) if "Price_num" in df.columns and not df.empty else None
    med_ppa   = float(df["Price_per_Acre"].dropna().median()) if "Price_per_Acre" in df.columns and not df.empty else None
    n_total   = int(len(df))
    n_acres   = int(df["Acres_num"].notna().sum()) if "Acres_num" in df.columns else 0
    pct_acres = (n_acres / n_total) if n_total > 0 else None

    out_df = df.copy()
    for col in ["Price_num","Acres_num"]:
        if col in out_df.columns: out_df = out_df.drop(columns=[col])

    if os.path.exists(book_path):
        wb = load_workbook(book_path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames and len(wb.sheetnames)==1: wb.remove(wb["Sheet"])
    ws = wb.create_sheet(sheet_name[:31])

    # header area
    ws["A1"].value="Stato";  ws["A1"].font=Font(bold=True)
    ws["A2"].value="Contea"; ws["A2"].font=Font(bold=True)
    ws["B1"].value=state_full_name(stato)
    ws["B2"].value=str(contea) if contea is not None else ""
    ws["C1"].value="Media $/Acre";   ws["C1"].font=Font(bold=True)
    ws["C2"].value="Mediana $/Acre"; ws["C2"].font=Font(bold=True)
    ws["D1"].value=avg_ppa if avg_ppa is not None else None
    ws["D2"].value=med_ppa if med_ppa is not None else None
    if ws["D1"].value is not None: ws["D1"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    if ws["D2"].value is not None: ws["D2"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws["E1"].value="media prezzi";   ws["E1"].font=Font(bold=True)
    ws["F1"].value=avg_price if avg_price is not None else None
    if ws["F1"].value is not None: ws["F1"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    # align + fill
    center = Alignment(horizontal="center")
    for col in ["A","B","C"]:
        for row in range(1, 3):
            ws[f"{col}{row}"].alignment = center
    ws["D1"].alignment = Alignment(horizontal="left")
    ws["D2"].alignment = Alignment(horizontal="left")
    ws["E1"].alignment = Alignment(horizontal="right")
    ws["F1"].alignment = Alignment(horizontal="left"); ws["F1"].font=Font(bold=True)

    fill = PatternFill(start_color="00EEEEEE", end_color="00EEEEEE", fill_type="solid")
    for r in (1,2):
        for c in ("A","B","C","D","E","F"):
            ws[f"{c}{r}"].fill = fill

    # table headers
    headers = list(out_df.columns)
    header_row = 4
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = Font(bold=True)

    # find special columns
    try:
        ppa_col_idx = headers.index("Price_per_Acre")+1
    except ValueError:
        ppa_col_idx = None
    try:
        link_col_idx = headers.index("Link")+1
    except ValueError:
        link_col_idx = None

    # data rows + hyperlink styling
    hyperlink_font = Font(underline="single", color="0000EE")
    for r_idx, row in enumerate(out_df.itertuples(index=False), start=header_row+1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            # currency for $/acre
            if ppa_col_idx is not None and c_idx == ppa_col_idx and isinstance(val, (int, float)):
                from openpyxl.styles import numbers
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            # hyperlink on Link column
            if link_col_idx is not None and c_idx == link_col_idx and isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.style = "Hyperlink"
                cell.font = hyperlink_font

    # bottom coverage
    from openpyxl.styles import Side, Border
    thin = Side(style="thin", color="999999")
    last = ws.max_row + 1
    ws.cell(row=last, column=1, value="Coverage Acres")
    cov_text = f"{n_acres}/{n_total}" + (f" ({pct_acres*100:.1f}%)" if pct_acres is not None else "")
    try:
        price_col_idx = headers.index("Price")+1
    except ValueError:
        price_col_idx = len(headers)
    ws.cell(row=last, column=price_col_idx, value=cov_text)
    for c in range(1, ws.max_column+1):
        ws.cell(row=last, column=c).border = Border(top=thin)

    _auto_fit(ws)
    wb.save(book_path)
    return avg_price, avg_ppa

def choose_output_path() -> str:
    """
    Politica: UN SOLO FILE fisso per ogni run.
    - Cancella (se possibile) 'risultati_zillow_media.xlsx'
    - Rimuove eventuali file timestampati 'risultati_zillow_media_*.xlsx'
    - Restituisce sempre OUT_BASE
    """
    # 1) elimina il fisso se esiste
    if os.path.exists(OUT_BASE):
        try:
            os.remove(OUT_BASE)
            log("[RESET] removed old", OUT_BASE)
        except Exception as e:
            log("[WARN] cannot remove", OUT_BASE, "->", e)

    # 2) pulizia di eventuali vecchi timestampati
    ts_files = glob.glob(os.path.join(os.path.dirname(__file__), "risultati_zillow_media_*.xlsx"))
    for pth in ts_files:
        try:
            os.remove(pth)
            log("[CLEAN] removed old timestamped:", os.path.basename(pth))
        except Exception as e:
            log("[WARN] cannot remove", pth, "->", e)

    log("[INFO] using fixed output:", OUT_BASE)
    return OUT_BASE

def main():
    log("[RUNNER] avvio")
    if len(sys.argv) < 2:
        log("Uso: python zillow_avg_runner.py zillow_gui_filters.json")
        sys.exit(2)
    cfg_path = sys.argv[1]
    log("[CFG] file:", cfg_path)
    with open(cfg_path, "r", encoding="utf-8") as f:
        cfg = json.load(f)

    p = cfg["params"]
    contea = str(p.get("Contea") or "").strip()
    stato  = str(p.get("Stato") or "").strip()
    region_id = p.get("Region ID")
    north = p.get("north"); south = p.get("south"); east = p.get("east"); west = p.get("west")

    vendita = cfg["vendita"]
    periods = cfg["periods"]
    min_lot = cfg.get("min_lot"); max_lot = cfg.get("max_lot")

    log("[PARAMS]", contea, stato, region_id, north, south, east, west)
    log("[FILTERS] vendita:", vendita, "periods:", periods, "lot:", min_lot, max_lot)

    # Output unico per tutto il run (sovrascritto)
    output_path = choose_output_path()
    log("[OUTPUT] selected:", output_path)

    summaries = []
    for tipo in vendita:
        for per in periods:
            tipo_path = "sold" if "sold" in str(tipo).lower() else "land"
            url = build_url(contea, stato, region_id, north, south, east, west, per, min_lot, max_lot, tipo_path)
            sheet = f"{tipo.replace(' ','_')}_{per}"
            log("[RUN]", sheet, "URL:", url)
            try:
                rows = zts.scrape(url)
                log("[OK] scrape rows:", len(rows))
            except Exception as e:
                log("[ERR] durante scrape:", e)
                log(traceback.format_exc())
                continue
            df = df_from_rows(rows)
            avg_price, avg_ppa = append_sheet_with_avg(output_path, sheet, df, stato, contea)
            summaries.append((sheet, len(df), avg_price, avg_ppa))
            log("[SAVED]", output_path, "sheet:", sheet, "rows:", len(df), "avg_price:", avg_price, "avg_ppa:", avg_ppa)

    log("[DONE] Output:", output_path)
    for s, n, ap, aa in summaries:
        log(f" - {s}: {n} righe | media prezzo: {ap if ap is not None else 'N/A'} | media $/Acre: {aa if aa is not None else 'N/A'}")
    return 0

if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as e:
        log("[FATAL]", e)
        log(traceback.format_exc())
        sys.exit(1)
