
import json
import re
import sys
import time
import gc
from dataclasses import dataclass, asdict
from typing import List, Optional

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.workbook.defined_name import DefinedName
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import undetected_chromedriver as uc

TEST_URL = "https://www.zillow.com/appling-county-ga/land/?searchQueryState=%7B%22pagination%22%3A%7B%7D%2C%22isMapVisible%22%3Atrue%2C%22mapBounds%22%3A%7B%22west%22%3A-83.10302324414062%2C%22east%22%3A-81.49627275585937%2C%22south%22%3A31.276637324224254%2C%22north%22%3A32.15744225314186%7D%2C%22regionSelection%22%3A%5B%7B%22regionId%22%3A1516%2C%22regionType%22%3A4%7D%5D%2C%22filterState%22%3A%7B%22sort%22%3A%7B%22value%22%3A%22globalrelevanceex%22%7D%2C%22sf%22%3A%7B%22value%22%3Afalse%7D%2C%22tow%22%3A%7B%22value%22%3Afalse%7D%2C%22mf%22%3A%7B%22value%22%3Afalse%7D%2C%22con%22%3A%7B%22value%22%3Afalse%7D%2C%22apa%22%3A%7B%22value%22%3Afalse%7D%2C%22manu%22%3A%7B%22value%22%3Afalse%7D%2C%22apco%22%3A%7B%22value%22%3Afalse%7D%2C%22lot%22%3A%7B%22min%22%3A0%2C%22max%22%3A87120%2C%22units%22%3Anull%7D%2C%22doz%22%3A%7B%22value%22%3A%2212m%22%7D%7D%2C%22isListVisible%22%3Atrue%2C%22usersSearchTerm%22%3A%22Appling%20County%20GA%22%7D"

OUTPUT_XLSX = "zillow_appling_test.xlsx"

@dataclass
class Row:
    price: Optional[str]
    acres: Optional[str]
    location: Optional[str]
    link: Optional[str]

def _to_float(x) -> Optional[float]:
    try:
        s = re.sub(r"[^0-9.\-]", "", str(x))
        return float(s) if s else None
    except Exception:
        return None

def extract_next_data(html: str):
    m = re.search(r'__NEXT_DATA__"\s*type="application/json">(.+?)</script>', html, re.DOTALL)
    if not m:
        return None
    try:
        return json.loads(m.group(1))
    except Exception:
        return None

def parse_location(address: Optional[str]) -> Optional[str]:
    if not address:
        return None
    m = re.search(r",\s*(.+)$", address)
    if m:
        return m.group(1).strip()
    return address.strip()

def _fmt_price_usd(val: Optional[float]) -> Optional[str]:
    if val is None:
        return None
    try:
        v = float(val)
    except Exception:
        return None
    return f"${v:,.0f}"

def _extract_numeric_price(it) -> Optional[float]:
    """
    Prova a leggere un prezzo 'numerico' da diverse chiavi che Zillow usa,
    specialmente per i risultati SOLD.
    """
    # livello top
    for k in ("unformattedPrice",):
        v = it.get(k)
        if isinstance(v, (int, float)):
            return float(v)

    # dentro hdpData.homeInfo
    home = (it.get("hdpData") or {}).get("homeInfo") or {}
    for k in ("price","unformattedPrice","soldPrice","soldPriceHigh","soldPriceLow","priceForHDP"):
        v = home.get(k)
        if isinstance(v, (int, float)):
            return float(v)

    # a volte zestimate è presente (fallback molto blando, usalo solo se niente altro)
    v = home.get("zestimate")
    if isinstance(v, (int, float)):
        return float(v)

    return None

def collect_rows_from_payload(payload) -> List[Row]:
    def _probe(root, path):
        cur = root
        for k in path:
            if isinstance(cur, dict) and k in cur:
                cur = cur[k]
            else:
                return None
        return cur

    buckets = [
        ["props", "pageProps", "searchPageState", "cat1", "searchResults", "listResults"],
        ["props", "pageProps", "searchPageState", "cat2", "searchResults", "listResults"],
    ]

    list_results = None
    for path in buckets:
        list_results = _probe(payload, path)
        if isinstance(list_results, list) and list_results:
            break

    out: List[Row] = []
    if not isinstance(list_results, list):
        return out

    for it in list_results:
        # PRICE
        price_num = _extract_numeric_price(it)
        price = _fmt_price_usd(price_num) if price_num is not None else (it.get("price") or (it.get("variableData") or {}).get("text"))

        # ACRES
        acres = None
        las = it.get("lotAreaString")
        if isinstance(las, str):
            m = re.search(r"([\d.,]+)\s*acres?", las, re.I)
            if m:
                acres = m.group(1)
        if acres is None:
            lot_value = it.get("lotArea") or (it.get("hdpData") or {}).get("homeInfo", {}).get("lotAreaValue")
            lot_unit = it.get("lotAreaUnit") or (it.get("hdpData") or {}).get("homeInfo", {}).get("lotAreaUnit")
            if lot_value is not None and isinstance(lot_unit, str) and lot_unit.lower().startswith("acre"):
                acres = str(lot_value)

        location = parse_location(it.get("address"))
        detail_url = it.get("detailUrl")
        if isinstance(detail_url, str) and detail_url.startswith("/"):
            link = "https://www.zillow.com" + detail_url
        else:
            link = detail_url

        out.append(Row(price=price, acres=acres, location=location, link=link))
    return out

def collect_rows_via_cards(driver) -> List[Row]:
    out: List[Row] = []
    cards = driver.find_elements(By.XPATH, "//article|//div[contains(@data-test,'search-list-item')]")
    for c in cards[:200]:
        txt = c.text
        # price
        m_price = re.search(r"\$\s?[\d,\.]+", txt)
        price = m_price.group(0) if m_price else None
        # acres
        m_acres = re.search(r"([\d,\.]+)\s*acres?", txt, re.I)
        acres = m_acres.group(1) if m_acres else None
        # location
        loc = None
        lines = [t.strip() for t in txt.splitlines() if t.strip()]
        for line in lines:
            if re.search(r",[ ]*[A-Z]{2}[ ]*\d{5}", line):
                loc = re.sub(r"^.*?,\s*", "", line)
                break
        # link
        try:
            a = c.find_element(By.XPATH, ".//a[@href]")
            href = a.get_attribute("href")
        except Exception:
            href = None
        out.append(Row(price=price, acres=acres, location=loc, link=href))
    return out

def scrape(url: str) -> List[Row]:
    # Monkeypatch to suppress noisy __del__ WinError 6 on teardown
    try:
        uc.Chrome.__del__ = lambda self: None  # type: ignore
    except Exception:
        pass

    driver = None
    try:
        chrome_options = uc.ChromeOptions()
        chrome_options.add_argument("--no-first-run")
        chrome_options.add_argument("--no-service-autorun")
        chrome_options.add_argument("--password-store=basic")
        # chrome_options.add_argument("--headless=new")  # opzionale

        driver = uc.Chrome(options=chrome_options)
        driver.get(url)

        try:
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//script[@id='__NEXT_DATA__']"))
            )
        except Exception:
            time.sleep(6)

        html = driver.page_source
        payload = extract_next_data(html)
        rows = collect_rows_from_payload(payload) if payload else []

        if not rows:
            rows = collect_rows_via_cards(driver)

        rows = [r for r in rows if (r.price or r.acres)]
        return rows
    finally:
        try:
            if driver is not None:
                driver.quit()
        except Exception:
            pass
        driver = None
        gc.collect()

def write_excel(rows: List[Row], out_path: str):
    # Build DataFrame with numeric helpers
    data = []
    for r in rows:
        price_num = None
        if r.price:
            price_num = re.sub(r"[^0-9.\-]", "", r.price)
            price_num = float(price_num) if price_num else None
        acres_num = float(r.acres.replace(",", "")) if r.acres else None
        data.append({
            "Price": r.price,
            "Price_num": price_num,
            "Acres": r.acres,
            "Acres_num": acres_num,
            "Location": r.location,
            "Link": r.link,
        })
    df = pd.DataFrame(data, columns=["Price", "Price_num", "Acres", "Acres_num", "Location", "Link"])
    # Write excel: single sheet + average row on same sheet
    # 1) compute average using Price_num if present
    avg_price = float(df["Price_num"].dropna().mean()) if not df.empty and "Price_num" in df.columns else None

    # 2) drop Price_num (duplicate of Price) before writing
    if "Price_num" in df.columns:
        df = df.drop(columns=["Price_num"])

    # 3) write data
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Foglio1", index=False)

    # 4) reopen and append average row
    wb = load_workbook(out_path)
    ws = wb["Foglio1"]

    # find column index for 'Price' to place the average
    headers = {cell.value: cell.column for cell in ws[1]}
    price_col_idx = headers.get("Price", ws.max_column)

    row = ws.max_row + 1
    ws.cell(row=row, column=1, value="Media prezzo (USD)")
    if avg_price is not None:
        cell = ws.cell(row=row, column=price_col_idx, value=avg_price)
        try:
            from openpyxl.styles import numbers
            cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        except Exception:
            pass
    wb.save(out_path)

    return len(df), avg_price

def main():
    print("[INFO] Carico pagina Zillow di test...")
    rows = scrape(TEST_URL)
    print(f"[INFO] Trovati {len(rows)} risultati utili.")
    n, avg = write_excel(rows, OUTPUT_XLSX)
    if avg is not None:
        print(f"[OK] Salvato '{OUTPUT_XLSX}' (righe: {n}). Media valore: {avg:,.2f} USD (Summary!B1, nome: media_valore)")
    else:
        print(f"[OK] Salvato '{OUTPUT_XLSX}' (righe: {n}). Nessun prezzo numerico per calcolare la media.")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("[ERRORE]", type(e).__name__, str(e))
        sys.exit(1)
