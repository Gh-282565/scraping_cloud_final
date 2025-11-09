# -*- coding: utf-8 -*-
"""
scraper_core/scraper.py
Versione antecedente FIXATA:
- CREA DUE FILE SEPARATI (Realtor/Zillow) con fogli ForSale/Sold
- Ordine colonne + rimozione Title/Sold date
- Auto-fit colonne
- Ritorna la LISTA dei file creati (non la cartella)
"""

import os
from datetime import datetime
from typing import List, Tuple, Optional
import pandas as pd

try:
    from . import realtor_scrape
    print("[IMPORT] realtor_scrape OK ->", getattr(realtor_scrape, "__file__", "?"))
except Exception as e:
    print("[IMPORT][ERR] realtor_scrape:", e)
    realtor_scrape = None

try:
    from . import zillow_scrape
except Exception:
    zillow_scrape = None


# -----------------------------------------------------
# Utility
# -----------------------------------------------------

def _ensure_results_dir(base_dir: str) -> str:
    results = os.path.join(base_dir, "results")
    os.makedirs(results, exist_ok=True)
    return results

def _now_tag() -> str:
    return datetime.utcnow().strftime("%Y%m%d_%H%M%S")

def _to_df(obj) -> Optional[pd.DataFrame]:
    if obj is None:
        return None
    if isinstance(obj, tuple) and len(obj) >= 1:
        obj = obj[0]
    if isinstance(obj, pd.DataFrame):
        return obj
    if isinstance(obj, list) and obj and isinstance(obj[0], dict):
        return pd.DataFrame(obj)
    if isinstance(obj, dict):
        return pd.DataFrame([obj])
    return None

def _normalize(df: pd.DataFrame, source: str) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    df = df.copy()
    if "Source" not in df.columns:
        df.insert(0, "Source", source)
    return df

def _summary_vals(df: pd.DataFrame) -> tuple[float, float, float]:
    if df is None or df.empty:
        return (0.0, 0.0, 0.0)
    p = pd.to_numeric(df.get("Price"), errors="coerce")
    ppa = pd.to_numeric(df.get("Price_per_Acre"), errors="coerce")
    return (
        float(p.mean(skipna=True)) if not p.dropna().empty else 0.0,
        float(ppa.mean(skipna=True)) if not ppa.dropna().empty else 0.0,
        float(ppa.median(skipna=True)) if not ppa.dropna().empty else 0.0,
    )

# === SPEC ordine colonne + drop indesiderate ===
COLUMN_ORDER = [
    "Stato", "Contea",
    "Population 2022", "Population 2023", "Var % Population",
    "% Owner occupied housing units",
    "Median Household Income 2023",
    "Median House or Condo Value 2023",
    "% Resident living Poverty 2023",
    "Unemployment in November 2024",
    "Median resident age",
    "Single-family new house construction building permits: 2023",
    "Contea Crime per 100k abit.2023",
    "Stato Crime per 100k abit.2023",
    "% Crime Contea Vs Stato 2023",
    "Indicazione Rischio Geologico", "Valore Rischio Geologico",
    "Land For Sale 30gg", "Land SOLD 30gg", "% Land SOLD Vs For Sale 30gg",
    "Land For Sale 90gg", "Land SOLD 90gg", "% Land SOLD Vs For Sale 90gg",
    "Land For Sale 6M",  "Land SOLD 6M",  "% Land SOLD Vs For Sale 6M",
    "Land For Sale 12M", "Land SOLD 12M", "% Land SOLD Vs For Sale 12M",
    # campi listing tipici (se presenti)
    "Source", "Status", "Period", "State", "County", "Location",
    "Price", "Acres", "Price_per_Acre", "Link"
]

UNWANTED_COLUMNS = {
    "Title",
    "Sold date", "Sold_date", "SoldDate",
    "Sale Date", "Sale_Date", "SaleDate"
}

def reorder_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    keep = [c for c in COLUMN_ORDER if c in df.columns]
    extra = [c for c in df.columns if c not in keep and c not in UNWANTED_COLUMNS]
    return df[keep + extra]

def drop_unwanted(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    return df.drop(columns=[c for c in df.columns if str(c) in UNWANTED_COLUMNS], errors="ignore")


# -----------------------------------------------------
# Excel save function (formattazione + ordine colonne)
# -----------------------------------------------------

def _save_excel(df_all: pd.DataFrame, outpath: str, source: str):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font

    # Split per ForSale/Sold
    df_forsale = df_all[df_all["Status"].str.contains("for sale", case=False, na=False)]
    df_sold    = df_all[df_all["Status"].str.contains("sold", case=False, na=False)]

    # Drop & reorder per ciascun foglio
    if df_forsale is not None and not df_forsale.empty:
        df_forsale = reorder_columns(drop_unwanted(df_forsale))
    if df_sold is not None and not df_sold.empty:
        df_sold = reorder_columns(drop_unwanted(df_sold))

    def fmt_sheet(ws, df, avg_price, avg_ppa, med_ppa):
        start_row = 6  # header alla riga 6, dati dalla 7
        for c in ws[start_row]:
            c.font = Font(bold=True)
            c.alignment = Alignment(wrap_text=True, vertical="center")

        ws.freeze_panes = f"A{start_row+1}"
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(ws.max_column)}{ws.max_row}"

        # Mappa colonna
        colmap = {col: (df.columns.get_loc(col) + 1) for col in df.columns}

        def fmt(col, pattern):
            if col in colmap:
                for r in range(start_row+1, ws.max_row+1):
                    ws.cell(row=r, column=colmap[col]).number_format = pattern

        # Formati tipici
        fmt("Price", '"$"#,##0')
        fmt("Acres", '#,##0.00')
        fmt("Price_per_Acre", '"$"#,##0.00')

        # Hyperlink su Link
        if "Link" in colmap:
            c = colmap["Link"]
            for r in range(start_row+1, ws.max_row+1):
                cell = ws.cell(row=r, column=c)
                val = str(cell.value or "").strip()
                if val.startswith("http"):
                    cell.hyperlink = val
                    cell.style = "Hyperlink"

        # Auto-fit colonne (cap 60)
        def txtlen(v): return len(str(v)) if v is not None else 0
        for j, col in enumerate(df.columns, start=1):
            maxlen = max(
                txtlen(col),
                max((txtlen(ws.cell(row=r, column=j).value)
                     for r in range(start_row+1, ws.max_row+1)), default=0)
            )
            ws.column_dimensions[get_column_letter(j)].width = min(maxlen + 4, 60)

        # Summary in alto
        ws.cell(row=1, column=1, value="Media Prezzo").font = Font(bold=True)
        ws.cell(row=1, column=2, value=avg_price).number_format = '"$"#,##0'
        ws.cell(row=2, column=1, value="Media Prezzo/Acro").font = Font(bold=True)
        ws.cell(row=2, column=2, value=avg_ppa).number_format = '"$"#,##0.00'
        ws.cell(row=3, column=1, value="Mediana Prezzo/Acro").font = Font(bold=True)
        ws.cell(row=3, column=2, value=med_ppa).number_format = '"$"#,##0.00'

    # Crea workbook e scrive i due fogli
    if os.path.exists(outpath):
        os.remove(outpath)
    wb = Workbook()
    wb.remove(wb.active)

    with pd.ExcelWriter(outpath, engine="openpyxl") as writer:
        for name, df in [("ForSale", df_forsale), ("Sold", df_sold)]:
            if df is None or df.empty:
                pd.DataFrame(["Nessun risultato"]).to_excel(writer, sheet_name=name, index=False, header=False)
                continue

            avg_price, avg_ppa, med_ppa = _summary_vals(df)
            df.to_excel(writer, sheet_name=name, index=False, startrow=5)
            ws = writer.sheets[name]
            fmt_sheet(ws, df, avg_price, avg_ppa, med_ppa)

    print(f"[OK] File Excel creato per {source}: {outpath}")


# -----------------------------------------------------
# Funzione principale orchestratore
# -----------------------------------------------------

def run_scraping(
    *,
    state: str,
    county: str,
    acres_min: int,
    acres_max: int,
    include_forsale: bool,
    include_sold: bool,
    use_sources: List[str],
    headless: bool = True,
    period: Optional[str] = None,
) -> Tuple[List[str], List[str]]:
    """
    Esegue Realtor e/o Zillow e crea file separati.
    Ritorna: (lista_file_creati, messages)
    """
        print("[RUN][SOURCES]", use_sources)
    messages: List[str] = []
    produced_paths: List[str] = []

    base_dir = os.path.abspath(os.path.dirname(__file__))
    project_dir = os.path.dirname(base_dir)
    results_dir = _ensure_results_dir(project_dir)
    tag = _now_tag()

    kwargs = dict(
        state=state,
        county=county,
        acres_min=acres_min,
        acres_max=acres_max,
        include_forsale=include_forsale,
        include_sold=include_sold,
        headless=headless,
        period=period,
    )

    # Realtor
    if "realtor" in [s.lower() for s in (use_sources or [])] and realtor_scrape is not None:
        try:
            fn_r = getattr(realtor_scrape, "run_scrape", None) or getattr(realtor_scrape, "run", None)
            if not callable(fn_r):
                raise AttributeError("realtor_scrape non espone run_scrape/run.")
            df_r = _to_df(fn_r(**kwargs))
            if df_r is not None and not df_r.empty:
                df_r = _normalize(df_r, "Realtor")
                outpath_r = os.path.join(results_dir, f"realtor_risultati_estrazione_{tag}.xlsx")
                _save_excel(df_r, outpath_r, "Realtor")
                produced_paths.append(outpath_r)
                messages.append("[OK] File Realtor creato.")
            else:
                messages.append("[WARN] Nessun risultato Realtor.")
        except Exception as e:
            messages.append(f"[ERR] Realtor: {e}")

    # Zillow
    if "zillow" in [s.lower() for s in (use_sources or [])] and zillow_scrape is not None:
        try:
            fn_z = getattr(zillow_scrape, "run_scrape", None) or getattr(zillow_scrape, "run", None)
            if not callable(fn_z):
                raise AttributeError("zillow_scrape non espone run_scrape/run.")
            df_z = _to_df(fn_z(**kwargs))
            if df_z is not None and not df_z.empty:
                df_z = _normalize(df_z, "Zillow")
                outpath_z = os.path.join(results_dir, f"zillow_risultati_estrazione_{tag}.xlsx")
                _save_excel(df_z, outpath_z, "Zillow")
                produced_paths.append(outpath_z)
                messages.append("[OK] File Zillow creato.")
            else:
                messages.append("[WARN] Nessun risultato Zillow.")
        except Exception as e:
            messages.append(f"[ERR] Zillow: {e}")

    if not produced_paths:
        messages.append("[WARN] Nessun file generato.")
    return produced_paths, messages
