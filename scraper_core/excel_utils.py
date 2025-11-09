# scraper_core/excel_utils.py
import os
from datetime import datetime
import pandas as pd
from openpyxl.utils import get_column_letter

RESULTS_DIR = "/app/results"

def _ensure_dir(p):
    os.makedirs(p, exist_ok=True)

def _autoformat(ws):
    # freeze & autofilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    # larghezza colonne 19 come tuo standard
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 19

def _to_df(records):
    # normalizza in DF con colonne base
    if not records:
        return pd.DataFrame(columns=["Title", "Price", "Acres", "Link", "Status"])
    rows = []
    for r in records:
        rows.append({
            "Title":  r.get("title", ""),
            "Price":  r.get("price", ""),
            "Acres":  r.get("acres", None),
            "Link":   r.get("link", ""),
            "Status": r.get("status", ""),
        })
    return pd.DataFrame(rows)

def save_realtor_results(results: dict, filename: str = None):
    """
    results = { "for_sale": [..], "sold": [..] }
    Salva in /app/results un xlsx con due fogli (ForSale/Sold).
    Ritorna filepath.
    """
    _ensure_dir(RESULTS_DIR)
    if not filename:
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"realtor_results_{ts}.xlsx"
    out_path = os.path.join(RESULTS_DIR, filename)

    df_fs = _to_df(results.get("for_sale", []))
    df_sd = _to_df(results.get("sold", []))

    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        df_fs.to_excel(xw, sheet_name="ForSale", index=False)
        df_sd.to_excel(xw, sheet_name="Sold", index=False)

        for name in ["ForSale", "Sold"]:
            ws = xw.book[name]
            _autoformat(ws)
            # trasforma i Link in hyperlink cliccabili
            # (openpyxl accetta direttamente le url nelle celle; pandas le ha già inserite come testo)
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=4)  # colonna D = Link
                if cell.value:
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"

    return out_path
